"""
BTE Platform
Excel Renderer
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from .base_renderer import BaseRenderer

from ..models import Chapter


class ExcelRenderer(BaseRenderer):

    name = "excel"

    extension = ".xlsx"

    mime_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    def render(
        self,
        chapters: list[Chapter],
    ) -> bytes:

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Interpretation"

        row = 1

        for chapter in chapters:

            worksheet.cell(
                row=row,
                column=1,
                value=chapter.title,
            )

            row += 1

            for paragraph in chapter.paragraphs:

                worksheet.cell(
                    row=row,
                    column=1,
                    value=paragraph.title,
                )

                row += 1

                for sentence in paragraph.sentences:

                    worksheet.cell(
                        row=row,
                        column=2,
                        value=sentence.text,
                    )

                    row += 1

                row += 1

            row += 1

        buffer = BytesIO()

        workbook.save(buffer)

        return buffer.getvalue()
